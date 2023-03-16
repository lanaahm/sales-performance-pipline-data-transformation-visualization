import os
import openpyxl
import csv
import re
from datetime import datetime

# set path data
base_path = os.path.abspath(__file__ + "\\..\\")
source_path = f"{base_path}\\data\\Sample Data.xlsx"
raw_path = f"{base_path}\\data\\raw_data.csv"
bronze_path = f"{base_path}\\data\\bronze_data.csv"

def extract_new_raw_data():
    """
    Extract new raw data from the current source
    """

    # create a new folder if it doesn't exists
    os.makedirs(os.path.dirname(raw_path), exist_ok=True)

    # open xlsx file using xlrd
    wb = openpyxl.load_workbook(source_path)

    # select the first worksheet
    ws = wb.active

    # create csv file
    with open(raw_path, 'w', newline='', encoding='windows-1252') as csv_file:
        fieldnames = {k.value: re.sub("\\(.*?\\)|[^a-zA-Z_]", "", k.value).lower() for k in ws[1]}
        writer = csv.DictWriter(csv_file, fieldnames=fieldnames)
        writer.writerow(fieldnames)

        print("[Extract] First row example:", fieldnames)

        # loop through each row in the worksheet
        for row in ws.iter_rows(min_row=1):
            row_values = [cell.value for cell in row]
            row_dict = {fieldname: value for fieldname, value in zip(fieldnames, row_values)}
            writer.writerow(row_dict)


def transform_case(input_string):
    """
    Lowercase string fields
    """
    return input_string.lower()


def transform_integer(input_field):
    """
    Convert float to integer
    """

    # Convert the input field to a float first
    float_value = float(input_field)
    # Convert the float value to an integer
    return int(float_value)


def update_date(date_input):
    """
    Update date format from DD/MM/YYYY to YYYY-MM-DD
    """
    date_patterns = ["%d-%m-%Y", "%m-%d-%Y", "%Y-%d-%m", "%Y-%m-%d"]
    for pattern in date_patterns:
        try:
            current_format = datetime.strptime(date_input, pattern)
            new_format = current_format.strftime("%Y-%m-%d")
            return new_format
        except:
            pass
    return date_input


def transform_new_raw_data():
    """
    Apply all transformations for each row in the csv file
    """
    with open(raw_path, mode="r", encoding="windows-1252") as csv_file:
        # Read the new .csv ready to be processed
        reader = csv.DictReader(csv_file)
        row = next(reader)  # Get first row from reader

        # Initialize an empty list for our objects
        raw_objects = []

        fieldnames = {k: re.sub("\(.*?\)|[^a-zA-Z_]", "", k).lower() for k in row}
        # Write headers as first line
        raw_objects.append(fieldnames)

        for row in reader:
            # Apply transformations and save as object
            raw_objects.append(
                [
                    update_date(row["orderdate"]),
                    transform_case(row["region"]),
                    transform_case(row["person"]),
                    transform_case(row["item"]),
                    transform_integer(row["units"]),
                    row["unitcost"],
                    row["total"]
                ]
            )
        return raw_objects


def load_data(transformed_data):
    """
    Insert operation: add new csv data
    """
    # open the output CSV file
    with open(bronze_path, 'w', newline='') as output_file:
        # create a CSV writer object
        writer = csv.writer(output_file)

        # write the transformed data to the output file
        writer.writerows(transformed_data)


if __name__ == '__main__':
    print(f"[Extract] Start")
    print(f"[Extract] data from '{source_path}' to '{raw_path}'")
    extract_new_raw_data()
    print(f"[Extract] End")

    print("[Transform] Start")
    transform_data = transform_new_raw_data()
    print("[Transform] End")

    print("[Load] Start")
    print("[Load] Inserting new rows data")
    load_data(transform_data)
    print("[Load] End")

